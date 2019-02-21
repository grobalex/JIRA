import requests
import pprint
import getpass
import json
import re
import sys
from openpyxl import load_workbook

'''
List of TA's:
'''
user_ids = []

'''
Username and password for Jira user.
CONSTRAINT: Needs to be an Admin user.
'''
username = ""
pswd = ""

"""
Get the JIRA login credentials
Sets:
	- GLOBAL String username
	- GLOBAL String pswd
"""


def getLoginCreds():
    global username, pswd
    username = getpass.getpass('Username:')
    pswd = getpass.getpass('Password:')

# These should be all Jira permission levels that need to be changed.
jirapermissions = ['ADMINISTER_PROJECTS', 'BROWSE_PROJECTS', 'MANAGE_SPRINTS_PERMISSION', 'VIEW_DEV_TOOLS', 'VIEW_READONLY_WORKFLOW', 'CREATE_ISSUES', 'EDIT_ISSUES', 'TRANSITION_ISSUES', 'SCHEDULE_ISSUES', 'MOVE_ISSUES', 'ASSIGN_ISSUES', 'ASSIGNABLE_USER', 'RESOLVE_ISSUES', 'CLOSE_ISSUES', 'MODIFY_REPORTER', 'DELETE_ISSUES', 'LINK_ISSUES', 'SET_ISSUE_SECURITY',
                   'VIEW_VOTERS_AND_WATCHERS', 'MANAGE_WATCHERS', 'ADD_COMMENTS', 'EDIT_ALL_COMMENTS', 'EDIT_OWN_COMMENTS', 'DELETE_ALL_COMMENTS', 'DELETE_OWN_COMMENTS', 'CREATE_ATTACHMENTS', 'DELETE_ALL_ATTACHMENTS', 'DELETE_OWN_ATTACHMENTS', 'WORK_ON_ISSUES', 'EDIT_OWN_WORKLOGS', 'EDIT_ALL_WORKLOGS', 'DELETE_OWN_WORKLOGS', 'DELETE_ALL_WORKLOGS']

# Disabled TLS verification warning
requests.packages.urllib3.disable_warnings()

# Base URL for JIRA
baseapiurl = 'https://jira.ccs.neu.edu/rest/api/2/'

'''
Determines if the provided key conforms to the JIRA requirements
Arguments:
	- String Key
Return:
	- Boolean
'''


def special_match(strg, search=re.compile(r'([A-Z][A-Z]+)[0-9_]*').search):
    return bool(search(strg))

'''
Creates the specific JSON payload for permissionschemes required by the JIRA API
Arguments:
	- List JIRA_GROUP : A List of JIRA groups
	- List permissions : All the JIRA permissions that we would like to change
	- String permissionschemename : The Permission scheme we would like to apply these changes to
Return:
	- A JSON object in the format that the JIRA API expects
'''


def createJSONPayload(JIRA_GROUP, permissions, permissionschemename):
    temp = []
    for permission in permissions:
        for group in JIRA_GROUP:
            temp.append('{"holder": {"type": "group","parameter":"' +
                        group + '"},"permission": "' + permission + '"}')
    return '{"name":"' + permissionschemename + '", "description": "description" , "permissions": [' + (','.join(map(str, temp))) + ']}'

'''
Sends a request to the JIRA API
Arguments:
	- String URL
	- String typeOfRequest: one-of: POST, GET, PUT, DELETE
		- Throws ValueError if not one-of typeOfRequest
	- JSON data
Return:
	- Successful connection to JIRA API and preformed action with confirmation message
	- ERROR message returned from the JIRA API
'''


def APIrequest(url, typeOfRequest, data):
    if typeOfRequest == 'post':
        response = requests.post(url, verify=False, auth=(username, pswd), headers={
                                 'Accept': 'application/json', 'Content-Type': 'application/json'}, json=data)
    elif typeOfRequest == 'get':
        response = requests.get(url, verify=False, auth=(username, pswd), headers={
            'Accept': 'application/json'})
    elif typeOfRequest == 'put':
        response = requests.put(url, verify=False, auth=(username, pswd), headers={
                                'Accept': 'application/json', 'Content-Type': 'application/json'}, json=data)
    elif typeOfRequest == 'delete':
        response = requests.delete(url, verify=False, auth=(
            username, pswd), headers={'Accept': 'application/json'})
    elif typeOfRequest == 'post':
        response = requests.post(url, verify=False, auth=(
            username, pswd), headers={'Accept': 'application/json'})
    else:
        raise ValueError(
            'Type of Request {} is not valid'.format(typeOfRequest))
    try:
        response.raise_for_status()
    except requests.exceptions.HTTPError as e:
        print("Error: " + str(e))
    json_obj = response.json()
    if baseapiurl + 'permissionscheme' in url:
        # PERMISSION SCHEME ID
        return json_obj['id']

'''
Creates the JSON payload to send to APIrequest to create a JIRA project
Arguments:
	- String key
	- String name
	- String projectTypeKey must be either software or business
	- String projectTemplateKey must be one of JIRA template URLs
	- String description
	- String lead must be CCIS-ID
	- Int permissionSchemeKey
Return:
	- Successful connection to JIRA API and create project
	- ERROR message returned from the JIRA API
'''


def createProject(key, name, projectTypeKey, projectTemplateKey, description, lead, permissionSchemeKey):
    print(key, name, projectTypeKey, projectTemplateKey,
          description, lead, permissionSchemeKey)
    APIrequest(baseapiurl + 'project', 'post', {
        "key": key,
        "name": name,
        "projectTypeKey": projectTypeKey,
        "projectTemplateKey": projectTemplateKey,
        "description": description,
        "lead": lead,
        "assigneeType": "PROJECT_LEAD",
        "permissionScheme": permissionSchemeKey})

'''
Adds users to a given JIRA group
Arguments:
	- List of Strings users
	- String URL

Return:
	- Successful connection to JIRA API and users added
	- ERROR message returned from the JIRA API
'''


def addUsersToGroup(users, URL):
    for user in users:
        APIrequest(URL, 'post', {'name': user})

'''
Creates a complete JIRA workflow
Arguments:
	- String name
	- String key
	- String projectTypeKey : must be either software or business
	- String projectTemplateKey : must be one of JIRA template URLs
	- String description
	- String lead : must be CCIS-ID
	- List of Strings users or none if not set
	- String ldap or none if not set
Return:
	- Successful connection to JIRA API
	- ERROR message returned from the JIRA API
'''


def createCompleteProject(name, key, projectTypeKey, projectTemplateKey, description, lead, users, ldap):
    APIrequest(baseapiurl + 'group', 'post', {"name": name + '-group'})
    if ldap is not None and users is None:
        permissionSchemeKey = APIrequest(baseapiurl + 'permissionscheme', 'post', json.loads(
            createJSONPayload([ldap], jirapermissions, name + '-PS')))
    else:
        permissionSchemeKey = APIrequest(baseapiurl + 'permissionscheme', 'post', json.loads(
            createJSONPayload([name + '-group'], jirapermissions, name + '-PS')))
        # print(permissionSchemeKey)
        addUsersToGroup(users, baseapiurl +
                        'group/user?groupname=' + name + '-group')
    createProject(key, name, projectTypeKey, projectTemplateKey,
                  description, lead, permissionSchemeKey)


'''
Get all permissions associated with a permissions scheme
Arguments:
    - String ID
'''


def get_permission_scheme(ID):
    APIrequest(baseapiurl + 'permissionscheme/' +
               ID + '/permission', 'get', None)


'''
Prompts the user for their input to create a new JIRA project
Depending on if the user selects software or business,
it will prompt the user to select the appropriate JIRA project template
The user also has the choice of selecting an LDAP group
or create a new JIRA group with users they input
Arguments:
	- String name
	- String key
	- String projectTypeKey must be either software or business
	- String projectTemplateKey must be one of JIRA template URLs
	- String description
	- String lead must be CCIS-ID
	- If LDAP group:
		- String LDAP-group
	-Else:
		- List of Strings: CCIS-IDs
Return:
	- Successful connection to JIRA API and create project
	- ERROR message returned from the JIRA API
'''


def userInput():
    name = input('Input Project Name: ')
    keyrequirements = '''
	##################### KEY MUST ONLY CONTAIN: ####################
	- The first TWO characters must be a upper case letter
	- Letters must be [A-Z] and upper case
	- Only letters, numbers or the underscore character can be used
	'''
    print(keyrequirements)
    while True:
        key = input('Input Project Key: ')
        if special_match(key):
            break
        else:
            continue
    projectTypeKey = input(
        'Should this be a Software (s) or Business (b) project?')
    print(projectTypeKey)
    if projectTypeKey.lower() == 's':
        projectTypeKey = 'software'
        while True:
            projectTemplate = input(
                'What template should be used? Scrum (s) , Kanban (k) or Soft. Dev. (sd): ')
            projectTemplate = projectTemplate.lower()
            if projectTemplate == 's':
                projectTemplateKey = 'com.pyxis.greenhopper.jira:gh-scrum-template'
                break
            elif projectTemplate == 'k':
                projectTemplateKey = 'com.pyxis.greenhopper.jira:gh-kanban-template'
                break
            elif projectTemplate == 'sd':
                projectTemplateKey = 'com.pyxis.greenhopper.jira:basic-software-development-template'
                break
            else:
                continue
    elif projectTypeKey.lower() == 'b':
        projectTypeKey = 'business'
        while True:
            projectTemplate = input(
                'What template should be used? Project (p) , Task (t) or Process (ps) management ')
            projectTemplate = projectTemplate.lower()
            if projectTemplate == 'p':
                projectTemplateKey = 'com.atlassian.jira-core-project-templates:jira-core-project-management'
                break
            elif projectTemplate == 't':
                projectTemplateKey = 'com.atlassian.jira-core-project-templates:jira-core-task-management'
                break
            elif projectTemplate == 'ps':
                projectTemplateKey = 'com.atlassian.jira-core-project-templates:jira-core-process-management'
                break
            else:
                continue
    description = input('Input a project description: ')
    lead = input('Input the project leads CCIS ID: ')
    uses_ldap = input('Will this Project use an LDAP group? (y/n) ')
    if uses_ldap.lower() == 'y':
        ldap = input('Type in the LDAP group name: ')
        users = None
    else:
        users = input(
            'Input a list of users seperated by a comma: ').split(",")
        ldap = None
    createCompleteProject(name, key, projectTypeKey,
                          projectTemplateKey, description, lead, users, ldap)

'''
iterate over all rows in the excel file and return its contents
Argument:
	- Sheet Object: An excel sheet
Return:
	- Return all contents of the row in a list
'''


def iterateOverRows(sheet):
    rowlist = []
    counter = 0
    for row in sheet.iter_rows():
        templist = []
        if counter != 0:
            for cell in row:
                templist.append(cell.value)
            rowlist.append(templist)
        counter += 1
    return rowlist

'''
Given a list of all users and a project key, return only the specific users for that specific project
Arguments:
	- A List of List of all users [project_key, CCIS-ID]
	- String: project key
Return:
	- A list of all CCIS-IDs
'''


def getUser(list_of_all_users, project_Key):
    project_user_list = []
    for user in list_of_all_users:
        if user[0] == project_Key:
            project_user_list.append(user[1])
    return project_user_list

'''
Given a path to file, extracts all information from a xlsx file
Arguments:
	- String: Path to file
Return:
	- Successful connection to JIRA API and create project
'''


def fileHandler(PATH, FLAG):
    global user_ids
    try:
        wb2 = load_workbook(PATH)
    except:
        raise ValueError('No .xlsx file given')
    if wb2.sheetnames[0] == 'groups' and wb2.sheetnames[1] == 'persons':
        if FLAG:
            for project in iterateOverRows(wb2[wb2.sheetnames[0]]):
                user_list = getUser(iterateOverRows(
                    wb2[wb2.sheetnames[1]]), project[1])

                createCompleteProject(project[0], project[1], project[2], project[
                    3], project[4], user_list[0], user_list, None)
        else:
            for project in iterateOverRows(wb2[wb2.sheetnames[0]]):
                print(project[0] + "-group")
                addUsersToGroup(user_ids, baseapiurl +
                                'group/user?groupname=' + project[0] + "-group")


'''
Main method
'''


def main():
    global user_ids
    if len(sys.argv) > 1:
        if sys.argv[1] == '--input':
            getLoginCreds()
            userInput()
        elif sys.argv[1] == '--file':
            if len(sys.argv) == 3:
                getLoginCreds()
                fileHandler(sys.argv[2], True)
            else:
                print('Missing the path to the file')
        elif sys.argv[1] == '--add':
            if len(sys.argv) == 4:
                getLoginCreds()
                addUsersToGroup(sys.argv[3].split(
                    ","), baseapiurl + 'group/user?groupname=' + sys.argv[2])
            else:
                print('Incorrect format')
                print("--add' followed by 'group-name' and 'CCIS-ID' adds user to group")
        elif sys.argv[1] == '--addall':
            if len(sys.argv) == 3:
                getLoginCreds()
                user_ids = input(
                    "Comma Seperated List of TA's CCIS-IDS: ").split(',')
                fileHandler(sys.argv[2], False)
            else:
                print('Incorrect format')
                print("--addall' Adds users with 'CCIS-ID' to all groups in excel file")
    else:
        helptext = '''
##################### How to use this JIRA CLI ####################

                 ! YOU MUST BE A JIRA ADMIN USER !
Use option:
'--input' to manual input values to create a new project
'--file' followed by 'path_to_file' to create projects based on .xlsx
'--add' followed by 'group-name' and comma seperated list of 'CCIS-ID' to add user(s) to a JIRA group
'--addall' Adds single or mulitple users with 'CCIS-ID' to all groups in excel file
'''
        print(helptext)
        exit()

main()

# createCompleteProject('test3', 'T3', 'software','com.pyxis.greenhopper.jira:gh-scrum-template', 'test', 'grobalex',['grobalex'])

# create permissions scheme
# APIrequest(baseapiurl + 'group', 'post', {"name": "example-group1"})
# APIrequest(baseapiurl + 'permissionscheme', 'post', json.loads(createJSONPayload(["example-group1"], jirapermissions, permissionschemename)))
# APIrequest(baseapiurl + 'group/user?groupname=example-group', 'post', {'name' : 'grobalex'})

# change permissions scheme
# APIrequest(baseapiurl + 'permissionscheme/10910', 'put', json.loads(createJSONPayload(groups, jirapermissions, permissionschemename)))

# create group
# APIrequest(baseapiurl + 'group', 'post', {"name": "example-group"})

# add users to group
# APIrequest(baseapiurl + 'group/user?groupname=example-group', 'post', {'name' : 'grobalex'})

# delete users from group
# APIrequest(baseapiurl + 'group/user?groupname=example-group', 'delete', {'name' : 'grobalex'})
